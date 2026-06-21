from pathlib import Path
from typing import Protocol


class WrittenSheet(Protocol):
    output_path: str


class SheetWriter(Protocol):
    def __call__(
        self,
        body: str,
        title: str,
        source_file: str,
        word_count: int,
        output_dir: Path,
        source_type: str,
        vault_dir: Path | None = None,
        header_extras: dict[str, str] | None = None,
    ) -> WrittenSheet: ...


def _cell_to_md(value) -> str:
    if value is None:
        return ""
    return str(value).replace("|", r"\|").replace("\n", " ").strip()


def _trim_rows(rows: list[list[str]]) -> list[list[str]]:
    while rows and not any(rows[-1]):
        rows.pop()

    if not rows:
        return []

    last_col = 0
    for row in rows:
        for index, cell in enumerate(row, start=1):
            if cell:
                last_col = max(last_col, index)

    return [row[:last_col] for row in rows]


def _rows_to_md(rows: list[list[str]]) -> str:
    trimmed_rows = _trim_rows(rows)
    if not trimmed_rows:
        return ""

    col_count = max(len(row) for row in trimmed_rows)
    normalized = [row + [""] * (col_count - len(row)) for row in trimmed_rows]
    lines = ["| " + " | ".join(normalized[0]) + " |"]
    lines.append("| " + " | ".join(["---"] * col_count) + " |")
    for row in normalized[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def write_xlsx_sheets(
    path: str,
    output_dir: Path,
    vault_dir: Path | None,
    write_sheet: SheetWriter,
) -> tuple[int, int, list[str]]:
    from openpyxl import load_workbook

    source_path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_count = 0
    total_words = 0
    output_paths: list[str] = []

    try:
        for sheet in workbook.worksheets:
            rows = [
                [_cell_to_md(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            body = _rows_to_md(rows)
            if not body:
                continue

            title = f"{source_path.stem} {sheet.title}"
            word_count = len(body.split())
            result = write_sheet(
                body,
                title,
                source_path.name,
                word_count,
                output_dir,
                "spreadsheets",
                vault_dir,
                header_extras={"Sheet": sheet.title},
            )
            sheet_count += 1
            total_words += word_count
            output_paths.append(Path(result.output_path).name)
    finally:
        workbook.close()

    return sheet_count, total_words, output_paths
